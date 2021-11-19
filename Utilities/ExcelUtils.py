import openpyxl

def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    Sheet=workbook[SheetName]
    return (Sheet.max_row)


def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    Sheet=workbook[SheetName]
    return (Sheet.max_column)


def readData(file,SheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    Sheet=workbook[SheetName]
    return Sheet.cell(row=rownum,column=columnno).value



def WriteData(file,SheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    Sheet=workbook[SheetName]
    Sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)


